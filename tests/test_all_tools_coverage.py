"""Coverage-oriented smoke/integration tests for all built-in PDF tools."""
from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZipFile

import pikepdf
import pytest

from pdf_agent.tools._builtins.add_blank_pages import AddBlankPagesTool
from pdf_agent.tools._builtins.add_page_numbers import AddPageNumbersTool
from pdf_agent.tools._builtins.auto_rotate import AutoRotateTool
from pdf_agent.tools._builtins.barcode import BarcodeTool
from pdf_agent.tools._builtins.booklet import BookletTool
from pdf_agent.tools._builtins.compare import CompareTool
from pdf_agent.tools._builtins.compress import CompressTool
from pdf_agent.tools._builtins.crop import CropTool
from pdf_agent.tools._builtins.decrypt import DecryptTool
from pdf_agent.tools._builtins.delete import DeleteTool
from pdf_agent.tools._builtins.deskew import DeskewTool
from pdf_agent.tools._builtins.encrypt import EncryptTool
from pdf_agent.tools._builtins.extract import ExtractTool
from pdf_agent.tools._builtins.extract_attachments import ExtractAttachmentsTool
from pdf_agent.tools._builtins.extract_images import ExtractImagesTool
from pdf_agent.tools._builtins.flatten import FlattenTool
from pdf_agent.tools._builtins.form_fill import FormFillTool
from pdf_agent.tools._builtins.header_footer import HeaderFooterTool
from pdf_agent.tools._builtins.images_to_pdf import ImagesToPdfTool
from pdf_agent.tools._builtins.linearize import LinearizeTool
from pdf_agent.tools._builtins.merge import MergeTool
from pdf_agent.tools._builtins.metadata_info import MetadataInfoTool
from pdf_agent.tools._builtins.nup import NUpTool
from pdf_agent.tools._builtins.ocr import OcrTool
from pdf_agent.tools._builtins.office_to_pdf import OfficeToPdfTool
from pdf_agent.tools._builtins.page_border import PageBorderTool
from pdf_agent.tools._builtins.pages_to_zip import PagesToZipTool
from pdf_agent.tools._builtins.pdf_to_html import PdfToHtmlTool
from pdf_agent.tools._builtins.pdf_to_images import PdfToImagesTool
from pdf_agent.tools._builtins.pdf_to_markdown import PdfToMarkdownTool
from pdf_agent.tools._builtins.pdf_to_office import PdfToExcelTool, PdfToPptTool
from pdf_agent.tools._builtins.pdf_to_pdfa import PdfATool
from pdf_agent.tools._builtins.pdf_to_text import PdfToTextTool
from pdf_agent.tools._builtins.pdf_to_word import PdfToWordTool
from pdf_agent.tools._builtins.qr_code import QrCodeTool
from pdf_agent.tools._builtins.redact import RedactTool
from pdf_agent.tools._builtins.remove_blank_pages import RemoveBlankPagesTool
from pdf_agent.tools._builtins.remove_metadata import RemoveMetadataTool
from pdf_agent.tools._builtins.reorder import ReorderTool
from pdf_agent.tools._builtins.repair import RepairTool
from pdf_agent.tools._builtins.resize import ResizeTool
from pdf_agent.tools._builtins.reverse_pages import ReversePagesTool
from pdf_agent.tools._builtins.rotate import RotateTool
from pdf_agent.tools._builtins.set_metadata import SetMetadataTool
from pdf_agent.tools._builtins.signature import SignatureTool
from pdf_agent.tools._builtins.signature_info import SignatureInfoTool
from pdf_agent.tools._builtins.split import SplitTool
from pdf_agent.tools._builtins.stamp import StampTool
from pdf_agent.tools._builtins.tile_pages import TilePagesTool
from pdf_agent.tools._builtins.validate import ValidateTool
from pdf_agent.tools._builtins.watermark_image import WatermarkImageTool
from pdf_agent.tools._builtins.watermark_text import WatermarkTextTool
from pdf_agent.tools.registry import load_builtin_tools, registry


def _tool_dir(workdir: Path, name: str) -> Path:
    path = workdir / name
    path.mkdir(parents=True, exist_ok=True)
    return path


def _assert_outputs_exist(result) -> None:
    assert result.output_files
    assert all(path.exists() for path in result.output_files)


class TestAllBuiltinsCoverage:
    def test_registry_loads_exactly_53_tools(self):
        if len(registry) == 0:
            load_builtin_tools()

        names = {manifest["name"] for manifest in registry.list_manifests()}
        assert len(names) == 53
        assert names == {
            "add_blank_pages",
            "add_page_numbers",
            "auto_rotate",
            "barcode",
            "booklet",
            "compare",
            "compress",
            "crop",
            "decrypt",
            "delete",
            "deskew",
            "encrypt",
            "extract",
            "extract_attachments",
            "extract_images",
            "flatten",
            "form_fill",
            "header_footer",
            "images_to_pdf",
            "linearize",
            "merge",
            "metadata_info",
            "nup",
            "ocr",
            "office_to_pdf",
            "page_border",
            "pages_to_zip",
            "pdf_to_excel",
            "pdf_to_html",
            "pdf_to_images",
            "pdf_to_markdown",
            "pdf_to_pdfa",
            "pdf_to_ppt",
            "pdf_to_text",
            "pdf_to_word",
            "qr_code",
            "redact",
            "remove_blank_pages",
            "remove_metadata",
            "reorder",
            "repair",
            "resize",
            "reverse_pages",
            "rotate",
            "set_metadata",
            "signature",
            "signature_info",
            "split",
            "stamp",
            "tile_pages",
            "validate",
            "watermark_image",
            "watermark_text",
        }

    def test_merge_split_booklet_nup_and_tile(self, rendered_text_pdf: Path, workdir: Path):
        merge_result = MergeTool().run(
            [rendered_text_pdf, rendered_text_pdf],
            {"mode": "sequential"},
            _tool_dir(workdir, "merge_full"),
        )
        with pikepdf.open(merge_result.output_files[0]) as pdf:
            assert len(pdf.pages) == 4

        split_result = SplitTool().run(
            [rendered_text_pdf],
            {"mode": "each_page"},
            _tool_dir(workdir, "split_each"),
        )
        assert len(split_result.output_files) == 2
        assert any("第0001页" in path.name for path in split_result.output_files)

        grouped_split_result = SplitTool().run(
            [rendered_text_pdf],
            {"mode": "range", "page_groups": "1|2"},
            _tool_dir(workdir, "split_groups"),
        )
        assert len(grouped_split_result.output_files) == 2
        assert all("按范围拆分" in path.name for path in grouped_split_result.output_files)

        booklet_result = BookletTool().run(
            [rendered_text_pdf],
            {},
            _tool_dir(workdir, "booklet"),
        )
        _assert_outputs_exist(booklet_result)

        nup_result = NUpTool().run(
            [rendered_text_pdf],
            {"layout": "2-up", "paper_size": "A4", "orientation": "portrait"},
            _tool_dir(workdir, "nup"),
        )
        _assert_outputs_exist(nup_result)

        tile_result = TilePagesTool().run(
            [rendered_text_pdf, rendered_text_pdf],
            {"direction": "horizontal"},
            _tool_dir(workdir, "tile"),
        )
        _assert_outputs_exist(tile_result)

    def test_page_manipulation_tools(self, sample_pdf: Path, blank_mix_pdf: Path, workdir: Path):
        rotate_result = RotateTool().run(
            [sample_pdf],
            {"angle": 90, "page_range": "1"},
            _tool_dir(workdir, "rotate"),
        )
        _assert_outputs_exist(rotate_result)

        blank_result = AddBlankPagesTool().run(
            [sample_pdf],
            {"page_range": "1,3", "count": 1},
            _tool_dir(workdir, "add_blank"),
        )
        with pikepdf.open(blank_result.output_files[0]) as pdf:
            assert len(pdf.pages) == 7

        extract_result = ExtractTool().run(
            [sample_pdf],
            {"page_range": "1-2"},
            _tool_dir(workdir, "extract"),
        )
        with pikepdf.open(extract_result.output_files[0]) as pdf:
            assert len(pdf.pages) == 2

        delete_result = DeleteTool().run(
            [sample_pdf],
            {"page_range": "2-5"},
            _tool_dir(workdir, "delete"),
        )
        with pikepdf.open(delete_result.output_files[0]) as pdf:
            assert len(pdf.pages) == 1

        reorder_result = ReorderTool().run(
            [sample_pdf],
            {"order": "5,4,3,2,1"},
            _tool_dir(workdir, "reorder"),
        )
        _assert_outputs_exist(reorder_result)

        reverse_result = ReversePagesTool().run(
            [sample_pdf],
            {},
            _tool_dir(workdir, "reverse"),
        )
        _assert_outputs_exist(reverse_result)

        crop_result = CropTool().run(
            [sample_pdf],
            {"top": 10, "bottom": 10, "left": 10, "right": 10, "page_range": "1"},
            _tool_dir(workdir, "crop"),
        )
        _assert_outputs_exist(crop_result)

        resize_result = ResizeTool().run(
            [sample_pdf],
            {"target_size": "A5", "page_range": "1-2"},
            _tool_dir(workdir, "resize"),
        )
        _assert_outputs_exist(resize_result)

        flatten_result = FlattenTool().run(
            [sample_pdf],
            {},
            _tool_dir(workdir, "flatten"),
        )
        _assert_outputs_exist(flatten_result)

        remove_blank_result = RemoveBlankPagesTool().run(
            [blank_mix_pdf],
            {"threshold": 0.98},
            _tool_dir(workdir, "remove_blank"),
        )
        with pikepdf.open(remove_blank_result.output_files[0]) as pdf:
            assert len(pdf.pages) == 2

    def test_annotation_and_watermark_tools(
        self,
        sample_pdf: Path,
        sample_images: list[Path],
        workdir: Path,
    ):
        text_wm_result = WatermarkTextTool().run(
            [sample_pdf],
            {"text": "张三专用", "font_size": 32, "page_range": "1"},
            _tool_dir(workdir, "wm_text"),
        )
        _assert_outputs_exist(text_wm_result)

        image_wm_result = WatermarkImageTool().run(
            [sample_pdf, sample_images[0]],
            {"opacity": 0.4, "scale": 0.2, "position": "center", "page_range": "1"},
            _tool_dir(workdir, "wm_image"),
        )
        _assert_outputs_exist(image_wm_result)

        stamp_result = StampTool().run(
            [sample_pdf, sample_images[1]],
            {"page_range": "1", "position": "bottom-right", "opacity": 0.7, "scale": 0.2},
            _tool_dir(workdir, "stamp"),
        )
        _assert_outputs_exist(stamp_result)

        header_footer_result = HeaderFooterTool().run(
            [sample_pdf],
            {"header": "Header", "footer": "Page {page}/{total}", "page_range": "1-2"},
            _tool_dir(workdir, "header_footer"),
        )
        _assert_outputs_exist(header_footer_result)

        page_numbers_result = AddPageNumbersTool().run(
            [sample_pdf],
            {"position": "bottom_center", "start_num": 10, "page_range": "1-3"},
            _tool_dir(workdir, "page_numbers"),
        )
        _assert_outputs_exist(page_numbers_result)

        border_result = PageBorderTool().run(
            [sample_pdf],
            {"border_width": 2, "border_color": "#FF0000", "page_range": "1"},
            _tool_dir(workdir, "page_border"),
        )
        _assert_outputs_exist(border_result)

        barcode_result = BarcodeTool().run(
            [sample_pdf],
            {"content": "ABC123", "barcode_type": "code128", "page_range": "1"},
            _tool_dir(workdir, "barcode"),
        )
        _assert_outputs_exist(barcode_result)

        qr_result = QrCodeTool().run(
            [sample_pdf],
            {"content": "https://example.com", "page_range": "1", "size": 60},
            _tool_dir(workdir, "qr"),
        )
        _assert_outputs_exist(qr_result)

        redact_result = RedactTool().run(
            [sample_pdf],
            {"regions_json": '[{"page":1,"x":50,"y":50,"width":80,"height":20}]', "fill_color": "black"},
            _tool_dir(workdir, "redact"),
        )
        _assert_outputs_exist(redact_result)

    def test_metadata_and_security_tools(
        self,
        sample_pdf: Path,
        encrypted_pdf: Path,
        signature_field_pdf: Path,
        sample_images: list[Path],
        form_pdf: Path,
        workdir: Path,
    ):
        metadata_result = MetadataInfoTool().run([sample_pdf], {}, _tool_dir(workdir, "metadata_info"))
        assert metadata_result.meta["page_count"] == 5

        set_meta_result = SetMetadataTool().run(
            [sample_pdf],
            {"title": "Hello", "author": "Tester"},
            _tool_dir(workdir, "set_meta"),
        )
        _assert_outputs_exist(set_meta_result)

        remove_meta_result = RemoveMetadataTool().run(
            [set_meta_result.output_files[0]],
            {},
            _tool_dir(workdir, "remove_meta"),
        )
        _assert_outputs_exist(remove_meta_result)

        encrypt_result = EncryptTool().run(
            [sample_pdf],
            {"owner_password": "owner-pass", "user_password": "user-pass"},
            _tool_dir(workdir, "encrypt_full"),
        )
        _assert_outputs_exist(encrypt_result)

        decrypt_result = DecryptTool().run(
            [encrypted_pdf],
            {"password": "userpass"},
            _tool_dir(workdir, "decrypt_full"),
        )
        _assert_outputs_exist(decrypt_result)

        signature_info_result = SignatureInfoTool().run(
            [signature_field_pdf],
            {},
            _tool_dir(workdir, "signature_info"),
        )
        assert signature_info_result.meta["has_signatures"] is True

        signature_result = SignatureTool().run(
            [sample_pdf, sample_images[0]],
            {"mode": "visible", "page": 1, "position": "bottom-right", "width_pt": 80},
            _tool_dir(workdir, "signature"),
        )
        _assert_outputs_exist(signature_result)

        form_fill_result = FormFillTool().run(
            [form_pdf],
            {"field_values": json.dumps({"name": "张三"}), "flatten": False},
            _tool_dir(workdir, "form_fill"),
        )
        _assert_outputs_exist(form_fill_result)

    def test_optimization_and_validation_tools(self, rendered_text_pdf: Path, workdir: Path):
        compress_result = CompressTool().run(
            [rendered_text_pdf],
            {"level": "high"},
            _tool_dir(workdir, "compress"),
        )
        _assert_outputs_exist(compress_result)

        repair_result = RepairTool().run(
            [rendered_text_pdf],
            {},
            _tool_dir(workdir, "repair"),
        )
        _assert_outputs_exist(repair_result)

        linearize_result = LinearizeTool().run(
            [rendered_text_pdf],
            {},
            _tool_dir(workdir, "linearize"),
        )
        _assert_outputs_exist(linearize_result)

        validate_result = ValidateTool().run(
            [rendered_text_pdf],
            {},
            _tool_dir(workdir, "validate"),
        )
        _assert_outputs_exist(validate_result)
        assert "is_valid" in validate_result.meta

        pdfa_result = PdfATool().run(
            [rendered_text_pdf],
            {"level": "2b"},
            _tool_dir(workdir, "pdfa"),
        )
        _assert_outputs_exist(pdfa_result)

    def test_image_and_attachment_extraction_tools(
        self,
        image_pdf: Path,
        attachment_pdf: Path,
        sample_images: list[Path],
        workdir: Path,
    ):
        images_to_pdf_result = ImagesToPdfTool().run(
            sample_images,
            {"page_size": "fit"},
            _tool_dir(workdir, "images_to_pdf"),
        )
        _assert_outputs_exist(images_to_pdf_result)

        pdf_to_images_result = PdfToImagesTool().run(
            [image_pdf],
            {"format": "png", "dpi": 96, "page_range": "all"},
            _tool_dir(workdir, "pdf_to_images"),
        )
        _assert_outputs_exist(pdf_to_images_result)
        assert pdf_to_images_result.output_files[0].suffix == ".png"

        pages_to_zip_result = PagesToZipTool().run(
            [image_pdf],
            {"format": "png", "dpi": 96},
            _tool_dir(workdir, "pages_to_zip"),
        )
        _assert_outputs_exist(pages_to_zip_result)
        with ZipFile(pages_to_zip_result.output_files[0]) as zf:
            assert zf.namelist()

        extract_images_result = ExtractImagesTool().run(
            [image_pdf],
            {},
            _tool_dir(workdir, "extract_images"),
        )
        _assert_outputs_exist(extract_images_result)

        extract_attachments_result = ExtractAttachmentsTool().run(
            [attachment_pdf],
            {},
            _tool_dir(workdir, "extract_attachments"),
        )
        _assert_outputs_exist(extract_attachments_result)
        assert extract_attachments_result.output_files[0].read_bytes() == b"attached hello"

    def test_text_and_markup_conversion_tools(self, rendered_text_pdf: Path, workdir: Path):
        text_result = PdfToTextTool().run(
            [rendered_text_pdf],
            {"page_range": "all"},
            _tool_dir(workdir, "pdf_to_text"),
        )
        _assert_outputs_exist(text_result)
        assert "Rendered page 1" in text_result.output_files[0].read_text(encoding="utf-8")

        html_result = PdfToHtmlTool().run(
            [rendered_text_pdf],
            {"single_page": True},
            _tool_dir(workdir, "pdf_to_html"),
        )
        _assert_outputs_exist(html_result)

        markdown_result = PdfToMarkdownTool().run(
            [rendered_text_pdf],
            {"preserve_layout": False},
            _tool_dir(workdir, "pdf_to_markdown"),
        )
        _assert_outputs_exist(markdown_result)
        assert "Page 1" in markdown_result.output_files[0].read_text(encoding="utf-8")

    def test_compare_tool(self, rendered_text_pdf: Path, workdir: Path):
        compare_result = CompareTool().run(
            [rendered_text_pdf, rendered_text_pdf],
            {"highlight_color": "red", "sensitivity": 10},
            _tool_dir(workdir, "compare"),
        )
        assert len(compare_result.output_files) == 2
        assert compare_result.output_files[0].suffix == ".pdf"
        assert compare_result.output_files[1].suffix == ".json"

    def test_ocr_tool(self, scanned_pdf: Path, workdir: Path):
        ocr_result = OcrTool().run(
            [scanned_pdf],
            {"language": "eng", "skip_text": True, "page_range": "1", "output_mode": "pdf"},
            _tool_dir(workdir, "ocr"),
        )
        _assert_outputs_exist(ocr_result)

    @pytest.mark.parametrize("fixture_name", ["sample_docx", "sample_xlsx", "sample_pptx"])
    def test_office_to_pdf_tool(self, fixture_name: str, workdir: Path, request: pytest.FixtureRequest):
        sample_file = request.getfixturevalue(fixture_name)
        result = OfficeToPdfTool().run(
            [sample_file],
            {},
            _tool_dir(workdir, "office_to_pdf"),
        )
        _assert_outputs_exist(result)
        assert result.output_files[0].suffix == ".pdf"
        assert result.meta["engine"] in {"libreoffice", "python-fallback"}
        assert result.meta["fallback_used"] == (result.meta["engine"] != "libreoffice")
        if result.meta["fallback_used"]:
            assert result.meta["fallback_reason"]
        assert "转PDF" in result.output_files[0].name

    @pytest.mark.parametrize(
        ("tool_cls", "suffix", "loader", "fallback_engine"),
        [
            (PdfToWordTool, ".docx", "docx", "python-docx"),
            (PdfToExcelTool, ".xlsx", "xlsx", "openpyxl"),
            (PdfToPptTool, ".pptx", "pptx", "python-pptx"),
        ],
    )
    def test_pdf_to_office_tools_real_output(
        self,
        rendered_text_pdf: Path,
        workdir: Path,
        tool_cls,
        suffix: str,
        loader: str,
        fallback_engine: str,
    ):
        result = tool_cls().run([rendered_text_pdf], {}, _tool_dir(workdir, f"office_{suffix[1:]}"))
        _assert_outputs_exist(result)
        assert result.output_files[0].suffix == suffix
        assert result.meta["engine"] in {"libreoffice", fallback_engine}
        assert result.meta["fallback_used"] == (result.meta["engine"] != "libreoffice")
        if result.meta["fallback_used"]:
            assert result.meta["fallback_reason"]
        if loader == "docx":
            from docx import Document

            doc = Document(result.output_files[0])
            assert doc.paragraphs
        elif loader == "xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(result.output_files[0])
            assert workbook.active.max_row >= 1
        else:
            from pptx import Presentation

            presentation = Presentation(result.output_files[0])
            assert len(presentation.slides) >= 1

    def test_auto_rotate_and_deskew_tools_real_execution(
        self,
        rotated_text_pdf: Path,
        workdir: Path,
    ):
        auto_rotate_result = AutoRotateTool().run(
            [rotated_text_pdf],
            {"min_confidence": 2},
            _tool_dir(workdir, "auto_rotate"),
        )
        _assert_outputs_exist(auto_rotate_result)
        assert auto_rotate_result.meta["rotations"]

        deskew_result = DeskewTool().run(
            [rotated_text_pdf],
            {"page_range": "1", "min_angle": 0.5},
            _tool_dir(workdir, "deskew"),
        )
        _assert_outputs_exist(deskew_result)
        assert deskew_result.meta["corrections"]
